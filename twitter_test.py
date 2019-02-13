from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time
from csv import writer
from openpyxl import load_workbook

# initialize test file
test_folder = 'C:\\dev\\twitter_test\\'
# initialize test folder
test_file = 'twitter_test.xlsx'

# logger initialize
ts = time.localtime()
logfile = 'Results' + time.strftime('%Y%m%d_%H%M%S', ts) + '.csv'
csvLogger = open(test_folder + logfile, 'w', newline='')
csvWriter = writer(csvLogger)
# write header for the logger
csvWriter.writerow(['TIMESTAMP', 'RESULT', 'STEP DETAILS'])
csvLogger.close()

def write_log(result, step_detail):
	csvLogger = open(test_folder + logfile, 'a', newline='')
	csvWriter = writer(csvLogger)
	csvWriter.writerow([time.strftime('%Y%m%d_%H%M%S', ts), result, step_detail])
	csvLogger.close()


# load test data into a dictionary
# initialize empty dictionary
dict_test_data = {}
# Read data from excel
wb = load_workbook(test_folder+test_file)
ws = wb['Test_Data']
for i in range(1, ws.max_column+1):
	dict_test_data[ws.cell(1, i).value] = ws.cell(2, i).value
wb.close()
write_log("DONE", "Test Data load complete")

# load test objects into a dictionary
# initialize empty dictionary
dict_test_obj = {}
# Read data from excel
wb = load_workbook(test_folder+test_file)
ws = wb['Test_Objects']
for i in range(2, ws.max_row+1):
	dict_test_obj[ws.cell(i, 1).value] = [ws.cell(i, 2).value, ws.cell(i, 3).value]
wb.close()
write_log("DONE", "Test Objects load complete")

# function to extract test data from dictionary


def get_value(test_data_name):
	return dict_test_data[test_data_name]

# function to extract test object from dictionary


def get_object(driver, test_object_name):
	if dict_test_obj[test_object_name][0] == 'NAME':
		return driver.find_element_by_name(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'ID':
		return driver.find_element_by_id(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'CSS_SELECTOR':
		return driver.find_element_by_css_selector(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'CLASS_NAME':
		return driver.find_element_by_class_name(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'XPATH':
		return driver.find_element_by_xpath(dict_test_obj[test_object_name][1])

# function to extract list of test objects from dictionary


def get_objects_list(driver, test_object_name):
	if dict_test_obj[test_object_name][0] == 'NAME':
		return driver.find_elements_by_name(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'ID':
		return driver.find_elements_by_id(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'CSS_SELECTOR':
		return driver.find_elements_by_css_selector(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'CLASS_NAME':
		return driver.find_elements_by_class_name(dict_test_obj[test_object_name][1])
	elif dict_test_obj[test_object_name][0] == 'XPATH':
		return driver.find_elements_by_xpath(dict_test_obj[test_object_name][1])

# function to initialize a driver - Chrome
# returns the webdriver object


def init_driver():
	# adding options to remove annoying pop up when chrome is opened via selenium
	options = webdriver.ChromeOptions()
	options.add_argument('--ignore-certificate-errors') 
	options.add_argument("--test-type")
	options.add_argument('--disable-extensions')
	options.add_argument('--start-maximized')
	options.add_experimental_option('useAutomationExtension', False)
	driver = webdriver.Chrome(options=options)
	driver.get('https://twitter.com')
	driver.implicitly_wait(5)
	# verify open
	try:
		get_object(driver, 'username')
	except NoSuchElementException:
		write_log('FAIL', 'Twitter login page load failed.')
	else:
		write_log('PASS', 'Twitter login page load success!')
	return driver

# function to login in to twitter
# return nothing


def login_twitter(driver, username, password):
	# enter username
	get_object(driver, 'username').send_keys(username)
	driver.implicitly_wait(1)
	# enter password
	get_object(driver, 'password').send_keys(password)
	driver.implicitly_wait(1)
	# login
	get_object(driver, 'password').submit()
	driver.implicitly_wait(5)
	# verify login
	try:
		get_object(driver, 'search_box')
	except NoSuchElementException:
		write_log('FAIL', 'Twitter login failed.')
	else:
		write_log('PASS', 'Twitter login success!')
	return

# function to search for a user in twitter
# return the object


def search_twitter(driver, search_string):	
	# search for the user specified input
	search_box = get_object(driver, 'search_box')
	search_box.clear()
	search_box.send_keys(search_string)
	# send a space to trigger the search
	search_box.send_keys(' ')
	time.sleep(3)

	# get the 1ist of search links and click for the matching handle
	twitter_handles = get_objects_list(driver, 'search_results')
	for handle in twitter_handles:
		# search for the user string in each twitter handle in the search results and click on it.
		if handle.get_attribute('data-user-screenname') == search_string:
			handle.click()
			driver.implicitly_wait(3)
			break


# function to validate the follow status and click follow


def test_follow_user(driver):
	# get the button to check follow status
	button_follow = get_object(driver, 'button_follow')

	# check follow
	if 'not-following' in button_follow.get_attribute('class'):
		# Account not followed. Follow ...
		write_log('INFO', 'Account not followed. Following...')
		button_follow.click()
		driver.implicitly_wait(3)
		# follow and log
		if 'not-following' not in button_follow.get_attribute('class'):
			write_log('PASS', 'Account followed Success!')
		else:
			write_log('FAIL', 'Account not followed.')
	else:
		# Account already followed
		write_log('PASS', 'Account already followed.')


# function to logout of twitter


def logout_twitter(driver):
	# logout
	get_object(driver, 'button_user').click()
	driver.implicitly_wait(3)
	get_object(driver, 'button_signout').click()
	driver.implicitly_wait(3)
	# verify logout
	try:
		get_object(driver, 'username')
	except NoSuchElementException:
		write_log('FAIL', 'Twitter signout failed.')
	else:
		write_log('PASS', 'Twitter signout success!')
	return

# function to close the driver/browser
# return nothing


def close_driver(driver):
	driver.close()
	return

# start


if __name__ == '__main__':

	# start a driver for a web browser:
	driver = init_driver()

	# get credentials from file
	username = get_value('UserName')
	password = get_value('Password')

	# log in to twitter
	login_twitter(driver, username, password)

	# get search string from file
	search_string = get_value('SearchString')

	# search twitter for particular string
	search_twitter(driver, search_string)

	# execute follow test case for search string
	test_follow_user(driver)

	# logout twitter
	logout_twitter(driver)

	# close the driver:
	close_driver(driver)
